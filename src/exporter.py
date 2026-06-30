import csv
import json
import os
from typing import List

from openpyxl import Workbook

from .schemas import DocumentTables, ExtractedTable


def export_json(doc_tables: DocumentTables, output_dir: str) -> str:
    """Export all tables to a single JSON file."""
    stem = os.path.splitext(doc_tables.source_file)[0]
    path = os.path.join(output_dir, f"{stem}_tables.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(doc_tables.model_dump(), f, indent=2, ensure_ascii=False)
    return path


def export_csv(doc_tables: DocumentTables, output_dir: str) -> List[str]:
    """Export each table as a separate CSV file."""
    stem = os.path.splitext(doc_tables.source_file)[0]
    paths = []
    for table in doc_tables.tables:
        path = os.path.join(output_dir, f"{stem}_{table.table_id}.csv")
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(table.columns)
            writer.writerows(table.rows)
        paths.append(path)
    return paths


def export_excel(doc_tables: DocumentTables, output_dir: str) -> str:
    """Export all tables to a single Excel workbook, one sheet per table."""
    stem = os.path.splitext(doc_tables.source_file)[0]
    path = os.path.join(output_dir, f"{stem}_tables.xlsx")
    wb = Workbook()
    wb.remove(wb.active)

    for table in doc_tables.tables:
        sheet_name = table.table_id[:31]
        ws = wb.create_sheet(title=sheet_name)
        ws.append(table.columns)
        for row in table.rows:
            ws.append(row)
        for col_idx in range(1, len(table.columns) + 1):
            ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = 18

    wb.save(path)
    return path
